import pandas as pd
import numpy as np
import ScraperFC as sfc
from datetime import datetime
from tqdm import tqdm
import warnings

from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Alignment

from typing import Union
from botasaurus.request import request, Request

ALL_LEAGUE = list(sfc.sofascore.comps.keys())
API_PREFIX = 'https://api.sofascore.com/api/v1'


from tqdm import tqdm
import pandas as pd

def get_league_matches(seasons_config='ALL', leagues_config='ALL'):
    """
    Retrieves match data for specified leagues and seasons from Sofascore.
    
    Parameters:
    -----------
    leagues_config : list or str
        League(s) to fetch match data for.
    seasons_config : list or str
        Season(s) to fetch match data for.
        
    Returns:
    --------
    pandas.DataFrame
        DataFrame containing match data sorted by startTimestamp in descending order.
    """
    # Initialize an empty DataFrame to store all matches
    all_matches = pd.DataFrame()

    # Convert single values into lists for consistent processing
    if not isinstance(leagues_config, list):
        leagues_config = [leagues_config]
    
    if not isinstance(seasons_config, list):
        seasons_config = [seasons_config]

    # Determine leagues to process
    if not leagues_config or leagues_config[0].upper() == "ALL":
        leagues = ALL_LEAGUE
    else:
        leagues = leagues_config

    for league in leagues:
        # Determine seasons for the current league
        if not seasons_config or seasons_config[0].upper() == "ALL":
            seasons = list(sfc.Sofascore().get_valid_seasons(league).keys())
        elif seasons_config[0].upper() == "LATEST":
            seasons = [list(sfc.Sofascore().get_valid_seasons(league).keys())[0]]
        else:
            seasons = seasons_config

        # Nested progress bar for seasons
        for season in tqdm(seasons, desc=f"Processing Seasons for `{league}`"):
            try:
                # Get match data from the Sofascore API
                matches = pd.DataFrame(
                    sfc.Sofascore().get_match_dicts(season, league)
                )
                
                # Add league and season columns for reference
                matches['league'] = league
                matches['season'] = season
                
                # Append to the main DataFrame
                all_matches = pd.concat([all_matches, matches], ignore_index=True)
                
            except Exception as e:
                print(f"Error fetching data for {league} in season {season}: {str(e)}")
    
    # Sort all matches by timestamp in descending order (most recent first)
    all_matches = all_matches.sort_values('startTimestamp', ascending=False)
    
    return all_matches

def get_teams_in_league(league, season):
    """Get all teams playing in the specified league season."""
    
    matches = get_league_matches(season, league)

    # Extract unique team names
    home_teams = matches['homeTeam'].apply(lambda x: x['name']).unique()
    away_teams = matches['awayTeam'].apply(lambda x: x['name']).unique()
    
    # Combine and get unique team names
    all_teams = np.unique(np.concatenate([home_teams, away_teams]))
    return all_teams

def get_team_matches(matches, team_name):
    """Get all matches for a specific team."""
    
    # Filter matches for the specified team (either home or away)
    team_matches = matches[(matches['homeTeam'].apply(lambda x: x['name']) == team_name) | 
                          (matches['awayTeam'].apply(lambda x: x['name']) == team_name)]
    
    return team_matches

def is_home_match(match, team_name):
    """Check if team is playing at home in this match."""
    return match['homeTeam']['name'] == team_name

@request(output=None, create_error_logs=False)
def scrape_odds(req: Request, match_id: Union[str, int]):
    """ Scrape odds data for a match
    
    Parameters
    ----------
    req : botasaurus.request.Request
        The request object provided by the botasaurus decorator
    match_id : str or int
        Sofascore match ID
        
    Returns
    -------
    : dict
        JSON response containing odds data
    """
    url = f'{API_PREFIX}/event/{match_id}/odds/1/all'
    response = req.get(url)
    
    if response.status_code == 200:
        odds_data = response.json()
        if not odds_data or 'markets' not in odds_data:
            print(f"**No odds data for match ID {match_id}")
            return pd.DataFrame()
        
        markets_list = []
        
        for market in odds_data['markets']:
            for choice in market.get('choices', []):
                markets_list.append({
                    'structureType': market.get('structureType'),
                    'marketId': market.get('marketId'),
                    'marketName': market.get('marketName'),
                    'isLive': market.get('isLive'),
                    'fid': market.get('fid'),
                    'suspended': market.get('suspended', False),
                    'id': market.get('id'),
                    'marketGroup': market.get('marketGroup'),
                    'marketPeriod': market.get('marketPeriod'),
                    'initialFractionalValue': choice.get('initialFractionalValue'),
                    'fractionalValue': choice.get('fractionalValue'),
                    'sourceId': choice.get('sourceId'),
                    'selection': choice.get('name'),
                    'winning': choice.get('winning'),
                    'change': choice.get('change'),
                    'choiceGroup': market.get('choiceGroup')
                })
        
        return pd.DataFrame(markets_list)
    else:
        # print(f"\nReturned {response.status_code} from {url}. Returning empty dictionary.")
        return pd.DataFrame()

def collect_match_data(match, team_name):
    """Collect match data for the specified team."""
    ss = sfc.Sofascore()
    match_id = int(match['id'])
    
    # Determine if the team is home or away
    is_home = is_home_match(match, team_name)
    # opponent_team = match['awayTeam']['name'] if is_home else match['homeTeam']['name']
    
    # Extract basic match info
    match_date = datetime.fromtimestamp(match['startTimestamp']).strftime('%d-%m-%Y')
    
    # Get scores
    team_score = match['homeScore']['current'] if is_home else match['awayScore']['current']
    opponent_score = match['awayScore']['current'] if is_home else match['homeScore']['current']
    
    # Get halftime scores
    ht_team_score = match['homeScore']['period1'] if is_home else match['awayScore']['period1']
    ht_opponent_score = match['awayScore']['period1'] if is_home else match['homeScore']['period1']
    
    # Determine W/L/D (W/L)
    if team_score > opponent_score:
        result = "W"
    elif team_score < opponent_score:
        result = "L"
    else:
        result = "D"

    # Total goals (Total of W)
    total_goals = team_score + opponent_score

    # Over/Under 2.5 goals (T/X)
    over_under = "T" if total_goals > 2.5 else "X"
        
    # Home or Away (Home/away)
    home_away = "H" if is_home else "A"
    
    # (Correct Score)
    correct_score = f"{team_score}-{opponent_score}"
    
    # Half Time Score (HT)
    ht_score = f"{ht_team_score}-{ht_opponent_score}"
    
    # Total Goal Odd/Even (Total Goal O/E)
    total_goal_oe = "O" if total_goals % 2 == 1 else "E"
    if total_goals == 0:
        total_goal_oe = "EN"
    
    # Team Score Odd/Even (Score O/E)
    team_score_oe = "O" if team_score % 2 == 1 else "E"
    if team_score == 0:
        team_score_oe = "EN"
    
    # Concede Odd/Even (Concede O/E)
    concede_oe = "O" if opponent_score % 2 == 1 else "E"
    if opponent_score == 0:
        concede_oe = "EN"
    
    # Both Teams To Score (BTTS)
    btts = "Y" if team_score > 0 and opponent_score > 0 else "N"

    # -----------Get match statistics------------

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        match_stats = ss.scrape_team_match_stats(match_id)

    # If no match stats are available, set to None
    corners_total = None
    cards_total = None
    corners_ht = None
    cards_ht = None
    
    if not match_stats.empty:
        corners = match_stats[match_stats['key'].str.contains("corner")]
        cards = match_stats[match_stats['key'].str.contains("Card")]
        
        try:
            # Extract totals of corners (Corner)
            corners_all = corners[corners['period'] == 'ALL'][['home', 'away']].iloc[0]
            corners_total = f"{corners_all['home']}-{corners_all['away']}" if is_home else f"{corners_all['away']}-{corners_all['home']}"
        except:
            pass

        try:
            # Extract totals of halftime-corners (Corner HT)
            corners_1st = corners[corners['period'] == '1ST'][['home', 'away']].iloc[0]
            corners_ht = f"{corners_1st['home']}-{corners_1st['away']}" if is_home else f"{corners_1st['away']}-{corners_1st['home']}"
        except:
            pass
            

        try:
            # Extract totals of cards (Card)
            cards_all = cards[cards['period'] == 'ALL'][['home', 'away']].astype(int).sum()
            cards_total = f"{cards_all['home']}-{cards_all['away']}" if is_home else f"{cards_all['away']}-{cards_all['home']}"
        except:
            pass
        
        try:
        # Extract totals of cards (Card HT)
            cards_1st = cards[cards['period'] == '1ST'][['home', 'away']].astype(int).sum()
            cards_ht = f"{cards_1st['home']}-{cards_1st['away']}" if is_home else f"{cards_1st['away']}-{cards_1st['home']}"
        except:
            pass

    # -----------Get match odds------------
    odds_data = scrape_odds(match_id)

    handicap_value = None
    
    try:
        selection = odds_data[odds_data['marketName'].str.contains('handicap')]['selection']
        selection = selection[selection.str.contains(team_name)].iloc[0]

        # Extract the handicap value (Handicap Value)
        handicap_value = float(selection.split(") ")[0].strip("("))
    except:
        pass

    # The asian handicap will be calculated based on the handicap value using the excel formula
    asian_handicap = None

    # Create a dictionary with all the processed data
    data = [
        result,
        over_under,
        total_goals,
        home_away,
        correct_score,
        ht_score,
        total_goal_oe,
        team_score_oe,
        concede_oe,
        btts,
        corners_total,
        cards_total,
        corners_ht,
        cards_ht,
        asian_handicap,
        handicap_value,
        match_date
    ]
    
    return data

def collect_teams_data(matches, team_names, columns):
    """Collect a dataframe with all match data for all teams in `team_names`."""
    team_data = {}
    
    for team_name in team_names:
        team_matches = get_team_matches(matches, team_name)
        
        # Process each match
        match_data = []
        
        for _, match in tqdm(team_matches.iterrows(), desc=f"Processing `{team_name}` matches", total=len(team_matches)):
            if match['status']['type'] == 'finished':  # Only completed matches
                try:
                    data = collect_match_data(match, team_name)

                    # Append the data to the match_data list
                    match_data.append(dict(zip(columns, data)))
                except Exception as e:
                    print(f"Error processing match `{int(match['id'])}` for `{team_name}`: `{str(e)}`")
                    continue
        
        # Create dataframe
        if match_data:
            team_data[team_name] = pd.DataFrame(match_data)
            print(f"Added data for `{team_name}` with `{len(match_data)}` rows")
        else:
            team_data[team_name] = pd.DataFrame()
            print(f'**Empty match data')
            
    return team_data

def save_excel(team_data, output_file="sofascore.xlsx"):
    """
    Create an Excel workbook with a sheet for each team.
    
    For each sheet:
      - The 'ASIAN' column receives a formula that determines results based on the 'Correct score' (column E)
        and 'Head Start' (column P).
      - Conditional formatting is applied for entire columns (rows 2 to 1,048,576):
          • 'W/L' and 'ASIAN': "L" gets a soft red, "W" a soft green, "D" light grey.
          • 'T/X': "X" gets a light red, "T" a light green.
          • 'BTTS': "N" gets a light red, "Y" a light green.
          • 'Total Goal O/E', 'Score O/E' and 'Concede O/E': "O" gets pastel blue, "E" pastel purple, "EN" soft magenta.
      - All cell values are centered, and the header row (row 1) is filled with a yellow color.
    """

    # Define soft colors for conditional formatting.
    # For 'W/L' and 'ASIAN'
    fill_L = PatternFill(fill_type="solid", start_color="F4CCCC", end_color="F4CCCC")   # soft red
    fill_W = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")   # soft green
    fill_D = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")   # light grey

    # For 'T/X'
    fill_X = PatternFill(fill_type="solid", start_color="FCE4D6", end_color="FCE4D6")   # light red
    fill_T = PatternFill(fill_type="solid", start_color="D9EAD3", end_color="D9EAD3")   # light green

    # For 'BTTS'
    fill_N = PatternFill(fill_type="solid", start_color="F8CBAD", end_color="F8CBAD")   # light red
    fill_Y = PatternFill(fill_type="solid", start_color="C9E2B3", end_color="C9E2B3")   # light green

    # For 'Total Goal O/E', 'Score O/E', 'Concede O/E'
    fill_O = PatternFill(fill_type="solid", start_color="BDD7EE", end_color="BDD7EE")   # pastel blue
    fill_E = PatternFill(fill_type="solid", start_color="E5D0FF", end_color="E5D0FF")   # pastel purple
    fill_EN = PatternFill(fill_type="solid", start_color="FCE4EC", end_color="FCE4EC")  # soft magenta

    # Header fill (yellow-ish)
    header_fill = PatternFill(fill_type="solid", start_color="FFEB3B", end_color="FFEB3B")

    # Create the Excel writer using openpyxl engine.
    writer = pd.ExcelWriter(output_file, engine='openpyxl')

    for team_name, match_data in team_data.items():
        if not match_data.empty:
            # Clean team name to create a valid sheet name (max 31 characters)
            sheet_name = team_name.replace("/", "-").replace("?", "")
            sheet_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name

            # Write the DataFrame to a new Excel sheet.
            match_data.to_excel(writer, sheet_name=sheet_name, index=False)

            worksheet = writer.sheets[sheet_name]
            # Determine the number of rows with data (header is row1)
            num_rows = match_data.shape[0]
            # We'll use the entire column (from row 2 to row 1,048,576) for conditional formatting.
            max_excel_row = 1048576
            last_col = worksheet.max_column

            # Insert the formula in the 'ASIAN' column (if present).
            if "ASIAN" in match_data.columns:
                asian_col_index = match_data.columns.get_loc("ASIAN") + 1  # Excel columns are 1-indexed.
                # Assuming "Correct score" is in column E and "Head Start" in column P.
                for row in range(2, num_rows + 2):
                    formula = (
                        f'=IF(VALUE(LEFT(E{row},FIND("-",E{row})-1))+P{row} > '
                        f'VALUE(MID(E{row},FIND("-",E{row})+1,LEN(E{row}))), "W", '
                        f'IF(VALUE(LEFT(E{row},FIND("-",E{row})-1))+P{row} = '
                        f'VALUE(MID(E{row},FIND("-",E{row})+1,LEN(E{row}))), "D", "L"))'
                    )
                    worksheet.cell(row=row, column=asian_col_index, value=formula)

            # Helper function to add conditional formatting rules
            # to the entire column (from row2 to row 1,048,576).
            def add_rules(column_name, rules):
                if column_name in match_data.columns:
                    col_idx = match_data.columns.get_loc(column_name) + 1
                    col_letter = get_column_letter(col_idx)
                    cell_range = f"{col_letter}2:{col_letter}{max_excel_row}"
                    for rule in rules:
                        worksheet.conditional_formatting.add(cell_range, rule)

            # Conditional formatting rules:
            # 1. For 'W/L' and 'ASIAN': "L", "W", "D"
            wl_rules = [
                CellIsRule(operator='equal', formula=['"L"'], fill=fill_L),
                CellIsRule(operator='equal', formula=['"W"'], fill=fill_W),
                CellIsRule(operator='equal', formula=['"D"'], fill=fill_D)
            ]
            add_rules("W/L", wl_rules)
            add_rules("ASIAN", wl_rules)

            # 2. For 'T/X': "X" and "T"
            tx_rules = [
                CellIsRule(operator='equal', formula=['"X"'], fill=fill_X),
                CellIsRule(operator='equal', formula=['"T"'], fill=fill_T)
            ]
            add_rules("T/X", tx_rules)

            # 3. For 'BTTS': "N" and "Y"
            btts_rules = [
                CellIsRule(operator='equal', formula=['"N"'], fill=fill_N),
                CellIsRule(operator='equal', formula=['"Y"'], fill=fill_Y)
            ]
            add_rules("BTTS", btts_rules)

            # 4. For 'Total Goal O/E', 'Score O/E', 'Concede O/E': "O", "E", "EN"
            oe_rules = [
                CellIsRule(operator='equal', formula=['"O"'], fill=fill_O),
                CellIsRule(operator='equal', formula=['"E"'], fill=fill_E),
                CellIsRule(operator='equal', formula=['"EN"'], fill=fill_EN)
            ]
            for col in ["Total Goal O/E", "Score O/E", "Concede O/E"]:
                add_rules(col, oe_rules)

            # Center-align all cells and apply header formatting (yellow fill) for row 1.
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=last_col):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.row == 1:
                        cell.fill = header_fill

            # Set column widths to be more readable
            for col_num, column in enumerate(match_data.columns, 1):
                col_letter = get_column_letter(col_num)
                # Set a reasonable column width (adjust as needed)
                worksheet.column_dimensions[col_letter].width = max(len(column) + 2, 12)
        else:
            print(f'**Empty match data: {team_name}')

    writer.close()
    print(f"Excel file saved as `{output_file}`")